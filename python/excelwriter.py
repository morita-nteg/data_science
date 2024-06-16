
from openpyxl import Workbook


def ofunc1(wb, sheet, n):
    ws = wb[sheet]
    
    ws['A' + str(n)] = 'exp2'

    ws['A' + str(n+1)] = 'Index'
    ws['B' + str(n+1)] = '0'
    ws['C' + str(n+1)] = '1'
    ws['D' + str(n+1)] = '2'
    ws['E' + str(n+1)] = '3'
    ws['A' + str(n+2)] = 'Value'
    ws['B' + str(n+2)] = '0'
    ws['C' + str(n+2)] = '2'
    ws['D' + str(n+2)] = '4'
    ws['E' + str(n+2)] = '8'

def ofunc2(wb, sheet, n):
    ws = wb[sheet]
    
    ws['A' + str(n)] = 'log2'

    ws['A' + str(n+1)] = 'Index'
    ws['B' + str(n+1)] = '1'
    ws['C' + str(n+1)] = '2'
    ws['D' + str(n+1)] = '4'
    ws['E' + str(n+1)] = '8'
    ws['A' + str(n+2)] = 'Value'
    ws['B' + str(n+2)] = '0'
    ws['C' + str(n+2)] = '1'
    ws['D' + str(n+2)] = '2'
    ws['E' + str(n+2)] = '3'


def writer(wbname, ofuncs):
    wb = Workbook()
    n = 1
    for ofunc in ofuncs:
        ofunc(wb, 'Sheet', n)
        n = n + 4

    wb.save(wbname)

if __name__ == '__main__':
    writer('sample.xlsx', [ofunc1, ofunc2])